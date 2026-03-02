"""
============================================================
  Google Maps – "No Website" Restaurant & Café Finder
============================================================
LEGAL NOTICE:
  Scraping Google Maps may violate Google's Terms of Service
  (https://policies.google.com/terms). This tool is provided
  for EDUCATIONAL PURPOSES ONLY. Before running in production,
  ensure compliance with:
    • Google Maps ToS
    • Local data-protection laws (GDPR, CCPA, etc.)
    • The target website's robots.txt
  The author assumes no liability for misuse.
============================================================

Requirements:
    pip install playwright pandas openpyxl
    playwright install chromium
"""

import asyncio
import logging
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout

# ──────────────────────────────────────────────
# Logging setup
# ──────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ──────────────────────────────────────────────
# 1. User input
# ──────────────────────────────────────────────
def get_user_inputs() -> tuple[str, str]:
    """Prompt the user for country and city."""
    print("\n" + "=" * 55)
    print("  Google Maps – No-Website Business Finder")
    print("=" * 55)
    country = input("  Enter country : ").strip()
    city    = input("  Enter city    : ").strip()
    if not country or not city:
        log.error("Country and city cannot be empty.")
        sys.exit(1)
    return country, city


# ──────────────────────────────────────────────
# 2. Build Google Maps search URL
# ──────────────────────────────────────────────
def build_maps_url(city: str, country: str) -> str:
    """Return a Google Maps search URL for cafes & restaurants."""
    query = f"cafes and restaurants in {city}, {country}"
    encoded = query.replace(" ", "+")
    url = f"https://www.google.com/maps/search/{encoded}"
    log.info("Search URL: %s", url)
    return url


# ──────────────────────────────────────────────
# 3. Scroll the results panel to load everything
# ──────────────────────────────────────────────
async def scroll_results(page, max_scrolls: int = 60, pause: float = 1.5) -> None:
    """
    Scroll the left-hand results panel on Google Maps until
    no more results load or max_scrolls is reached.
    """
    log.info("Scrolling results panel …")

    # The results list container selector (stable as of 2024-25)
    panel_selector = 'div[role="feed"]'

    try:
        panel = await page.wait_for_selector(panel_selector, timeout=15_000)
    except PlaywrightTimeout:
        log.warning("Results panel not found – page structure may have changed.")
        return

    prev_count = 0
    for i in range(max_scrolls):
        await panel.evaluate("el => el.scrollBy(0, 800)")
        await asyncio.sleep(pause)

        # Count currently loaded result cards
        cards = await page.query_selector_all('a[href*="/maps/place/"]')
        current_count = len(cards)

        log.info("  Scroll %d/%d – %d results loaded", i + 1, max_scrolls, current_count)

        # Stop when the panel signals "end of list"
        end_marker = await page.query_selector('span.HlvSq')   # "You've reached the end"
        if end_marker:
            log.info("  Reached end of results list.")
            break

        # Stop when count hasn't grown for two consecutive scrolls
        if current_count == prev_count and i > 2:
            log.info("  No new results after scroll – stopping.")
            break
        prev_count = current_count


# ──────────────────────────────────────────────
# 4. Extract business details from a result card
# ──────────────────────────────────────────────
async def extract_business_details(page, card) -> dict | None:
    """
    Click a result card, wait for the detail panel, and
    extract name, address, and website (if any).
    Returns None on failure.
    """
    try:
        await card.click()
        # Wait for the detail panel heading to appear
        await page.wait_for_selector('h1.DUwDvf, h1[class*="fontHeadlineLarge"]', timeout=8_000)
        await asyncio.sleep(0.8)   # let dynamic content settle
    except PlaywrightTimeout:
        log.warning("  Detail panel did not load – skipping.")
        return None

    # ── Name ──────────────────────────────────
    try:
        name_el = await page.query_selector('h1.DUwDvf, h1[class*="fontHeadlineLarge"]')
        name    = (await name_el.inner_text()).strip() if name_el else "N/A"
    except Exception:
        name = "N/A"

    # ── Address ───────────────────────────────
    address = "N/A"
    try:
        # Address buttons have a data-item-id starting with "address"
        addr_el = await page.query_selector('button[data-item-id="address"]')
        if addr_el:
            address = (await addr_el.inner_text()).strip()
        else:
            # Fallback: look for the copy-address aria-label
            addr_el2 = await page.query_selector('[aria-label*="Address"]')
            if addr_el2:
                address = (await addr_el2.get_attribute("aria-label") or "N/A").replace("Address: ", "").strip()
    except Exception:
        pass

    # ── Website ───────────────────────────────
    website = None
    try:
        web_el = await page.query_selector('a[data-item-id="authority"]')
        if web_el:
            website = (await web_el.get_attribute("href") or "").strip()
    except Exception:
        pass

    return {"name": name, "address": address, "website": website}


# ──────────────────────────────────────────────
# 5. Main scraping orchestrator
# ──────────────────────────────────────────────
async def scrape_maps(city: str, country: str) -> list[dict]:
    """
    Launch Playwright, open Google Maps, scroll results,
    click each card, and collect businesses without a website.
    """
    url = build_maps_url(city, country)
    results: list[dict] = []

    async with async_playwright() as pw:
        log.info("Launching browser …")
        browser = await pw.chromium.launch(
            headless=True,          # Set False to watch the browser live
            args=["--no-sandbox", "--disable-dev-shm-usage"],
        )
        context = await browser.new_context(
            locale="en-US",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        page = await context.new_page()

        log.info("Navigating to Google Maps …")
        await page.goto(url, wait_until="domcontentloaded", timeout=30_000)

        # Dismiss cookie/consent dialog if present (EU regions)
        try:
            consent_btn = await page.wait_for_selector(
                'button[aria-label*="Accept"], form[action*="consent"] button',
                timeout=5_000,
            )
            await consent_btn.click()
            await asyncio.sleep(1)
            log.info("Cookie consent dismissed.")
        except PlaywrightTimeout:
            pass   # No consent dialog – fine

        # Scroll to load all results
        await scroll_results(page)

        # Collect all result card links
        cards = await page.query_selector_all('a[href*="/maps/place/"]')
        # De-duplicate by href
        seen_hrefs: set[str] = set()
        unique_cards = []
        for card in cards:
            href = await card.get_attribute("href") or ""
            if href not in seen_hrefs:
                seen_hrefs.add(href)
                unique_cards.append(card)

        total = len(unique_cards)
        log.info("Found %d unique business cards. Extracting details …", total)

        for idx, card in enumerate(unique_cards, start=1):
            log.info("[%d/%d] Processing …", idx, total)
            details = await extract_business_details(page, card)

            if details is None:
                continue

            # Filter: keep only businesses WITHOUT a website
            if not details["website"]:
                log.info("  ✓ No website found: %s", details["name"])
                results.append({
                    "Business Name": details["name"],
                    "Address":       details["address"],
                    "Website":       "No Website",
                })
            else:
                log.info("  ✗ Has website – skipping: %s", details["name"])

            # Polite delay between requests
            await asyncio.sleep(0.5)

        await browser.close()

    log.info("Scraping complete. %d businesses without a website found.", len(results))
    return results


# ──────────────────────────────────────────────
# 6. Export to Excel
# ──────────────────────────────────────────────
def export_to_excel(records: list[dict], city: str, country: str) -> Path:
    """
    Save the filtered records to a formatted .xlsx file.
    Returns the output file path.
    """
    # Sanitise city/country for use in filename
    safe_city    = re.sub(r"[^\w\-]", "_", city.lower())
    safe_country = re.sub(r"[^\w\-]", "_", country.lower())
    filename     = f"no_website_restaurants_{safe_city}_{safe_country}.xlsx"
    output_path  = Path(filename)

    if not records:
        log.warning("No records to export – Excel file will not be created.")
        return output_path

    df = pd.DataFrame(records, columns=["Business Name", "Address", "Website"])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="No Website Businesses")

        # ── Basic formatting ──────────────────
        ws = writer.sheets["No Website Businesses"]

        # Column widths
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 15

        # Header style
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2E7D32")   # dark green
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Freeze header row
        ws.freeze_panes = "A2"

    log.info("Excel file saved → %s", output_path.resolve())
    return output_path


# ──────────────────────────────────────────────
# 7. Entry point
# ──────────────────────────────────────────────
async def main() -> None:
    country, city = get_user_inputs()

    log.info("Starting scrape for: %s, %s", city, country)
    start = time.perf_counter()

    records = await scrape_maps(city, country)
    output  = export_to_excel(records, city, country)

    elapsed = time.perf_counter() - start
    print("\n" + "=" * 55)
    print(f"  Done in {elapsed:.1f}s")
    print(f"  Businesses without website : {len(records)}")
    print(f"  Output file                : {output.resolve()}")
    print("=" * 55 + "\n")


if __name__ == "__main__":
    asyncio.run(main())
