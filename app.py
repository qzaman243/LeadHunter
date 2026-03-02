"""
╔═══════════════════════════════════════════════════════╗
║     LeadHunter Pro – Web GUI  (Flask + SSE)           ║
║     Run: python app.py  →  http://localhost:5000       ║
╚═══════════════════════════════════════════════════════╝

pip install flask playwright pandas openpyxl stripe
playwright install chromium
"""

import asyncio
import json
import logging
import os
import re
import threading
import time
import uuid
from pathlib import Path
from queue import Queue, Empty

import stripe
from flask import (Flask, render_template_string, request,
                   jsonify, send_file, Response, session, redirect, url_for)

# ── Scraper engine (inline, same logic as v2) ─────────────────────
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)-8s  %(message)s")

# ── Stripe config (replace with your keys) ───────────────────────
stripe.api_key = os.getenv("STRIPE_SECRET_KEY", "sk_test_YOUR_KEY_HERE")
STRIPE_PUB_KEY = os.getenv("STRIPE_PUBLIC_KEY", "pk_test_YOUR_KEY_HERE")
PRICE_PER_NICHE = 1000  # $10.00 in cents

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "leadhunter-super-secret-2025")

# ── In-memory stores ─────────────────────────────────────────────
job_queues: dict[str, Queue] = {}  # job_id → log Queue
job_results: dict[str, list] = {}  # job_id → records list
job_status: dict[str, str] = {}  # job_id → running|done|error
paid_sessions: set[str] = set()  # session IDs that paid (demo)

# ─────────────────────────────────────────────────────────────────
NICHES: dict[str, list[str]] = {
    "Home & Repair Services": ["plumbers", "electricians", "HVAC repair", "roofers", "carpenters", "painters",
                               "appliance repair", "garage door repair", "pest control"],
    "Automotive Businesses": ["auto repair shops", "truck repair shops", "car detailing", "tire shops",
                              "towing services", "auto body shops", "car wash", "used car dealers"],
    "Hospitality & Nightlife": ["bars", "pubs", "lounges", "nightclubs", "shisha cafes", "hookah lounges",
                                "catering services", "event venues"],
    "Personal Services": ["barber shops", "beauty salons", "nail salons", "massage centers", "spas", "tattoo studios",
                          "personal trainers"],
    "Health & Medical": ["dentists", "chiropractors", "physiotherapists", "private clinics", "mental health counselors",
                         "optometrists"],
    "Fitness Businesses": ["gyms", "CrossFit studios", "yoga studios", "martial arts schools", "boxing gyms"],
    "Professional Services": ["lawyers", "accountants", "tax consultants", "real estate agents", "mortgage brokers",
                              "immigration consultants"],
    "Property & Construction": ["real estate agencies", "construction companies", "remodeling contractors",
                                "interior designers", "architecture firms"],
    "Retail Stores": ["furniture stores", "electronics shops", "hardware stores", "clothing boutiques",
                      "Islamic clothing shops", "perfume stores", "organic stores"],
    "Logistics & Transport": ["moving companies", "trucking companies", "freight brokers", "courier services",
                              "taxi services"],
}

NICHE_ICONS = {
    "Home & Repair Services": "🏗", "Automotive Businesses": "🚗",
    "Hospitality & Nightlife": "🍸", "Personal Services": "💇",
    "Health & Medical": "🏥", "Fitness Businesses": "🏋️",
    "Professional Services": "🏢", "Property & Construction": "🏠",
    "Retail Stores": "🛒", "Logistics & Transport": "🚚",
}

FREE_NICHES = ["Hospitality & Nightlife", "Retail Stores"]  # Free tier access


# ─────────────────────────────────────────────────────────────────
# Scraper engine
# ─────────────────────────────────────────────────────────────────
def classify_lead(reviews_str, rating_str):
    try:
        reviews = int(str(reviews_str).replace(",", ""))
    except:
        reviews = 0
    try:
        rating = float(rating_str)
    except:
        rating = 0.0
    if reviews >= 10 and rating >= 3.5:
        return "HIGH LEAD"
    elif reviews >= 1:
        return "MEDIUM LEAD"
    else:
        return "LOW LEAD"


async def run_scrape(city, country, selected_niches, job_id, single_minor=None):
    q = job_queues[job_id]
    records = []

    def emit(msg):
        q.put(msg)
        log.info(msg)

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True, args=["--no-sandbox"])
        ctx = await browser.new_context(locale="en-US",
                                        user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36")
        page = await ctx.new_page()

        for niche in selected_niches:
            # If a specific minor sub-type was chosen, scrape only that one
            subs = [single_minor] if single_minor else NICHES[niche]
            emit(f"📂 Category: {niche}  →  Sub-type: {subs[0]}")

            for sub in subs:
                query = f"{sub} in {city}, {country}".replace(" ", "+")
                url = f"https://www.google.com/maps/search/{query}"
                emit(f"  🔍 Searching: {sub}")
                try:
                    await page.goto(url, wait_until="domcontentloaded", timeout=30_000)
                    await asyncio.sleep(1.5)

                    try:
                        btn = await page.wait_for_selector(
                            'button[aria-label*="Accept"], form[action*="consent"] button', timeout=4_000)
                        await btn.click();
                        await asyncio.sleep(1)
                    except PlaywrightTimeout:
                        pass

                    # Scroll
                    try:
                        panel = await page.wait_for_selector('div[role="feed"]', timeout=12_000)
                        prev = 0;
                        stale = 0
                        for _ in range(60):
                            await panel.evaluate("el => el.scrollBy(0, 900)")
                            await asyncio.sleep(1.1)
                            cards = await page.query_selector_all('a[href*="/maps/place/"]')
                            cnt = len(cards)
                            end = await page.query_selector('span.HlvSq')
                            if end: break
                            stale = stale + 1 if cnt == prev else 0
                            if stale >= 3: break
                            prev = cnt
                    except PlaywrightTimeout:
                        pass

                    cards = await page.query_selector_all('a[href*="/maps/place/"]')
                    seen = set();
                    unique = []
                    for c in cards:
                        h = await c.get_attribute("href") or ""
                        if h not in seen: seen.add(h); unique.append(c)

                    emit(f"  📍 {len(unique)} listings found for '{sub}'")
                    found_here = 0

                    for card in unique:
                        try:
                            await card.click()
                            await page.wait_for_selector('h1.DUwDvf, h1[class*="fontHeadlineLarge"]', timeout=8_000)
                            await asyncio.sleep(0.8)
                        except PlaywrightTimeout:
                            continue

                        # ── Extract fields with multi-fallback selectors ──
                        d = {"name": "N/A", "address": "N/A", "phone": "N/A", "rating": "N/A", "reviews": "N/A",
                             "website": None}

                        # NAME
                        try:
                            el = await page.query_selector('h1.DUwDvf, h1[class*="fontHeadlineLarge"]')
                            if el: d["name"] = (await el.inner_text()).strip()
                        except:
                            pass

                        # ADDRESS — 5-layer fallback
                        try:
                            # Layer 1: data-item-id attribute (old stable selector)
                            el = await page.query_selector('button[data-item-id="address"]')
                            if el:
                                d["address"] = (await el.inner_text()).strip()
                            else:
                                # Layer 2: aria-label contains "Address"
                                el = await page.query_selector('[aria-label^="Address:"], [aria-label^="Address "]')
                                if el:
                                    raw = (await el.get_attribute("aria-label") or "").strip()
                                    d["address"] = re.sub(r"^Address[:\s]+", "", raw).strip()
                                else:
                                    # Layer 3: copy-address button
                                    el = await page.query_selector(
                                        'button[jsaction*="address"], button[jsaction*="copy-address"]')
                                    if el:
                                        d["address"] = (await el.inner_text()).strip()
                                    else:
                                        # Layer 4: scan all buttons for street-like text
                                        btns = await page.query_selector_all('button.CsEnBe, div.rogA2c button')
                                        for btn in btns:
                                            txt = (await btn.inner_text()).strip()
                                            # Matches patterns like "123 Main St" or "45 Road, City"
                                            if re.search(r'\d+\s+\w+.{5,}', txt) and len(txt) < 120:
                                                d["address"] = txt
                                                break
                                        else:
                                            # Layer 5: page full-text regex scan for address pattern
                                            content = await page.content()
                                            match = re.search(
                                                r'"(\d{1,5}\s[\w\s\.,#-]{5,60}(?:St|Ave|Rd|Blvd|Dr|Ln|Way|Court|Ct|Pl|Sq|Street|Avenue|Road|Boulevard|Drive|Lane)[^"]{0,40})"',
                                                content
                                            )
                                            if match:
                                                d["address"] = match.group(1).strip()
                        except:
                            pass

                        # PHONE — 5-layer fallback (with international number detection)
                        try:
                            # Layer 1: data-item-id starts with "phone"
                            el = await page.query_selector('button[data-item-id^="phone"]')
                            if el:
                                d["phone"] = (await el.inner_text()).strip()
                            else:
                                # Layer 2: aria-label contains "Phone"
                                el = await page.query_selector('[aria-label^="Phone:"], [aria-label^="Phone "]')
                                if el:
                                    raw = (await el.get_attribute("aria-label") or "").strip()
                                    d["phone"] = re.sub(r"^Phone[:\s]+", "", raw).strip()
                                else:
                                    # Layer 3: href="tel:..." link
                                    el = await page.query_selector('a[href^="tel:"]')
                                    if el:
                                        href = await el.get_attribute("href") or ""
                                        d["phone"] = href.replace("tel:", "").strip()
                                    else:
                                        # Layer 4: scan button text for phone number patterns
                                        # Matches: +1 212-555-0100, +44 20 7946 0958, (212) 555-0100, etc.
                                        btns = await page.query_selector_all('button, span[class*="Io6YTe"]')
                                        phone_pattern = re.compile(
                                            r'(\+?\d{1,3}[\s\-.])?'  # country code
                                            r'(\(?\d{2,4}\)?[\s\-.])'  # area code
                                            r'(\d{3,4}[\s\-.])'  # first part
                                            r'(\d{3,4})'  # second part
                                        )
                                        for btn in btns:
                                            try:
                                                txt = (await btn.inner_text()).strip()
                                                if phone_pattern.search(txt) and len(txt) < 30:
                                                    d["phone"] = txt
                                                    break
                                            except:
                                                continue
                                        else:
                                            # Layer 5: regex scan full page HTML for phone numbers
                                            content = await page.content()
                                            # Prioritise international format (+country_code ...)
                                            intl_match = re.search(
                                                r'(\+\d{1,3}[\s\-\.]?\(?\d{1,4}\)?[\s\-\.]?\d{3,4}[\s\-\.]?\d{3,4})',
                                                content
                                            )
                                            if intl_match:
                                                d["phone"] = intl_match.group(1).strip()
                                            else:
                                                local_match = re.search(
                                                    r'(\(?\d{3}\)?[\s\-\.]\d{3}[\s\-\.]\d{4})',
                                                    content
                                                )
                                                if local_match:
                                                    d["phone"] = local_match.group(1).strip()
                        except:
                            pass

                        # Clean up phone: normalise country-code prefix spacing
                        if d["phone"] and d["phone"] != "N/A":
                            # Ensure country code has + prefix if it looks like one
                            d["phone"] = re.sub(r'^00(\d)', r'+\1', d["phone"])  # 0044... → +44...
                            d["phone"] = d["phone"].strip()

                        # RATING
                        try:
                            el = await page.query_selector('div.F7nice span[aria-hidden="true"]')
                            if not el:
                                el = await page.query_selector('span[aria-label*="stars"], span[aria-label*="star"]')
                            if el:
                                raw = (await el.inner_text()).strip() or (await el.get_attribute("aria-label") or "")
                                nums = re.findall(r'\d+\.?\d*', raw)
                                if nums: d["rating"] = nums[0]
                        except:
                            pass

                        # REVIEWS
                        try:
                            el = await page.query_selector('div.F7nice span[aria-label*="review"]')
                            if not el:
                                el = await page.query_selector(
                                    'span[aria-label*="review"], button[aria-label*="review"]')
                            if el:
                                label = await el.get_attribute("aria-label") or ""
                                nums = re.findall(r"[\d,]+", label)
                                if nums: d["reviews"] = nums[0].replace(",", "")
                            else:
                                # Fallback: find text like "(1,234)" or "1,234 reviews" near rating
                                content = await page.inner_text('body') if await page.query_selector('body') else ""
                                rev_match = re.search(r'\((\d[\d,]*)\)', content)
                                if rev_match: d["reviews"] = rev_match.group(1).replace(",", "")
                        except:
                            pass

                        # WEBSITE
                        try:
                            el = await page.query_selector('a[data-item-id="authority"]')
                            if not el:
                                el = await page.query_selector(
                                    'a[href*="http"][aria-label*="website" i], a[jsaction*="website"]')
                            if el:
                                d["website"] = (await el.get_attribute("href") or "").strip()
                        except:
                            pass

                        if d["website"]: continue

                        lead = classify_lead(d["reviews"], d["rating"])
                        records.append({
                            "Niche": niche, "Sub-Type": sub.title(),
                            "Business Name": d["name"], "Address": d["address"],
                            "Phone": d["phone"], "Rating": d["rating"],
                            "Reviews": d["reviews"], "Website Status": "No Website",
                            "Lead Quality": lead,
                        })
                        found_here += 1
                        await asyncio.sleep(0.4)

                    emit(f"  ✅ {found_here} no-website leads found for '{sub}'")
                    await asyncio.sleep(0.8)

                except Exception as e:
                    emit(f"  ⚠️ Error on '{sub}': {e}")

        await browser.close()

    job_results[job_id] = records
    job_status[job_id] = "done"
    emit(f"🎉 DONE — {len(records)} total leads found!")


def export_excel(records, city, country):
    safe = lambda s: re.sub(r"[^\w\-]", "_", s.lower())
    path = Path(f"/tmp/leads_{safe(city)}_{safe(country)}_{int(time.time())}.xlsx")
    if not records: return path

    df = pd.DataFrame(records)
    lead_colors = {"HIGH LEAD": "1B5E20", "MEDIUM LEAD": "E65100", "LOW LEAD": "B71C1C"}
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="1A237E")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    col_widths = {"Niche": 28, "Sub-Type": 22, "Business Name": 35, "Address": 45,
                  "Phone": 18, "Rating": 8, "Reviews": 10, "Website Status": 14, "Lead Quality": 14}

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="All Leads")
        ws = writer.sheets["All Leads"]
        _style(ws, df, col_widths, hdr_font, hdr_fill, hdr_align, border, lead_colors)

        for niche in df["Niche"].unique():
            sub = df[df["Niche"] == niche].reset_index(drop=True)
            nm = re.sub(r"[^\w\s]", "", niche).strip()[:31]
            sub.to_excel(writer, index=False, sheet_name=nm)
            _style(writer.sheets[nm], sub, col_widths, hdr_font, hdr_fill, hdr_align, border, lead_colors)

        summary = (df.groupby(["Niche", "Lead Quality"]).size().unstack(fill_value=0).reset_index())
        summary["Total"] = summary.select_dtypes("number").sum(axis=1)
        summary.to_excel(writer, index=False, sheet_name="Summary")

    return path


def _style(ws, df, col_widths, hdr_font, hdr_fill, hdr_align, border, lead_colors):
    for cell in ws[1]:
        cell.font = hdr_font;
        cell.fill = hdr_fill
        cell.alignment = hdr_align;
        cell.border = border
    ws.freeze_panes = "A2";
    ws.row_dimensions[1].height = 30
    headers = [c.value for c in ws[1]]
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(h, 16)
    lci = headers.index("Lead Quality") + 1 if "Lead Quality" in headers else None
    for ri, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(name="Arial", size=9)
        if ri % 2 == 0:
            for cell in row: cell.fill = PatternFill("solid", fgColor="F5F5F5")
        if lci:
            lc = row[lci - 1];
            color = lead_colors.get(str(lc.value), "")
            if color:
                lc.fill = PatternFill("solid", fgColor=color)
                lc.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
                lc.alignment = Alignment(horizontal="center", vertical="center")


# ─────────────────────────────────────────────────────────────────
# Flask Routes
# ─────────────────────────────────────────────────────────────────
HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>LeadHunter Pro</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=JetBrains+Mono:wght@400;600&display=swap" rel="stylesheet">
<style>
:root{
  --bg:#050811;--surface:#0d1424;--card:#111827;
  --accent:#00ff88;--accent2:#ff6b35;--accent3:#7c3aed;
  --text:#e2e8f0;--muted:#64748b;--border:#1e2d45;
  --high:#22c55e;--med:#f97316;--low:#ef4444;
}
*{margin:0;padding:0;box-sizing:border-box}
body{background:var(--bg);color:var(--text);font-family:'Syne',sans-serif;min-height:100vh;overflow-x:hidden}

/* Animated grid background */
body::before{
  content:'';position:fixed;inset:0;
  background-image:linear-gradient(var(--border) 1px,transparent 1px),
    linear-gradient(90deg,var(--border) 1px,transparent 1px);
  background-size:40px 40px;opacity:.3;pointer-events:none;z-index:0
}

/* Glow orbs */
.orb{position:fixed;border-radius:50%;filter:blur(100px);pointer-events:none;z-index:0;opacity:.15}
.orb1{width:500px;height:500px;background:var(--accent3);top:-100px;left:-100px}
.orb2{width:400px;height:400px;background:#0891b2;bottom:-100px;right:-100px}

.container{max-width:1100px;margin:0 auto;padding:24px;position:relative;z-index:1}

/* Nav */
nav{display:flex;align-items:center;justify-content:space-between;padding:0 0 32px}
.logo{font-size:1.5rem;font-weight:800;letter-spacing:-.02em}
.logo span{color:var(--accent)}
.badge{background:var(--accent3);color:#fff;font-size:.65rem;padding:2px 8px;border-radius:20px;font-weight:600;letter-spacing:.08em;vertical-align:top}
.nav-links{display:flex;gap:24px;align-items:center}
.nav-links a{color:var(--muted);text-decoration:none;font-size:.85rem;transition:.2s}
.nav-links a:hover{color:var(--text)}

/* Hero */
.hero{text-align:center;padding:40px 0 48px}
.hero h1{font-size:clamp(2rem,5vw,3.5rem);font-weight:800;line-height:1.1;letter-spacing:-.03em;margin-bottom:16px}
.hero h1 .green{color:var(--accent)}
.hero h1 .orange{color:var(--accent2)}
.hero p{color:var(--muted);font-size:1rem;max-width:520px;margin:0 auto 32px;line-height:1.7}
.plan-pills{display:flex;gap:12px;justify-content:center;flex-wrap:wrap;margin-bottom:8px}
.pill{padding:6px 16px;border-radius:99px;font-size:.75rem;font-weight:600;letter-spacing:.05em}
.pill-free{background:#0a2a1a;color:var(--accent);border:1px solid var(--accent)}
.pill-pro{background:#2d1a00;color:var(--accent2);border:1px solid var(--accent2)}

/* Pricing cards */
.pricing{display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:40px}
.price-card{background:var(--card);border:1px solid var(--border);border-radius:16px;padding:28px;position:relative;overflow:hidden;transition:.3s}
.price-card:hover{transform:translateY(-2px);border-color:var(--muted)}
.price-card.featured{border-color:var(--accent2)}
.price-card.featured::before{content:'MOST POPULAR';position:absolute;top:12px;right:-24px;background:var(--accent2);color:#fff;font-size:.6rem;padding:4px 32px;transform:rotate(35deg);font-weight:700;letter-spacing:.1em}
.price-card h3{font-size:1rem;font-weight:700;margin-bottom:8px}
.price-amount{font-size:2rem;font-weight:800;margin-bottom:4px}
.price-amount small{font-size:.9rem;color:var(--muted);font-weight:400}
.price-desc{color:var(--muted);font-size:.8rem;margin-bottom:20px;line-height:1.6}
.price-features{list-style:none;font-size:.82rem;color:var(--muted);display:flex;flex-direction:column;gap:8px}
.price-features li::before{content:'✓ ';color:var(--accent);font-weight:700}
.price-features li.no::before{content:'✗ ';color:var(--low)}

/* Main form card */
.card{background:var(--card);border:1px solid var(--border);border-radius:20px;padding:32px;margin-bottom:24px}
.section-label{font-size:.7rem;font-weight:700;letter-spacing:.15em;color:var(--accent);text-transform:uppercase;margin-bottom:20px}
.form-row{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:24px}
.form-group label{display:block;font-size:.75rem;color:var(--muted);margin-bottom:6px;font-weight:600;letter-spacing:.05em;text-transform:uppercase}
.form-group input{width:100%;background:var(--surface);border:1px solid var(--border);color:var(--text);padding:12px 16px;border-radius:10px;font-size:.9rem;font-family:'Syne',sans-serif;outline:none;transition:.2s}
.form-group input:focus{border-color:var(--accent);box-shadow:0 0 0 3px rgba(0,255,136,.08)}

/* Niche grid */
.niche-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:12px;margin-bottom:28px}
.niche-item{position:relative}
.niche-item input[type=checkbox]{position:absolute;opacity:0;width:0;height:0}
.niche-label{display:flex;align-items:center;gap:10px;padding:14px 16px;background:var(--surface);border:1px solid var(--border);border-radius:12px;cursor:pointer;transition:.2s;font-size:.82rem;user-select:none}
.niche-label:hover{border-color:var(--muted)}
.niche-item input:checked + .niche-label{border-color:var(--accent);background:rgba(0,255,136,.06);color:var(--accent)}
.niche-label .icon{font-size:1.2rem;flex-shrink:0}
.niche-label .name{font-weight:600;line-height:1.3}
.niche-label .locked{margin-left:auto;font-size:.65rem;background:var(--accent2);color:#fff;padding:2px 6px;border-radius:4px;font-weight:700}
.niche-label .free-tag{margin-left:auto;font-size:.65rem;background:var(--accent);color:#000;padding:2px 6px;border-radius:4px;font-weight:700}

/* Buttons */
.btn{display:inline-flex;align-items:center;gap:8px;padding:14px 28px;border-radius:12px;font-size:.9rem;font-weight:700;font-family:'Syne',sans-serif;cursor:pointer;border:none;transition:.2s;letter-spacing:.02em}
.btn-primary{background:var(--accent);color:#000}
.btn-primary:hover{background:#00e67a;transform:translateY(-1px)}
.btn-pay{background:var(--accent2);color:#fff}
.btn-pay:hover{background:#e55a25;transform:translateY(-1px)}
.btn-full{width:100%;justify-content:center}
.btn:disabled{opacity:.4;cursor:not-allowed;transform:none}

/* Payment banner */
.pay-banner{background:linear-gradient(135deg,#1a0a00,#2d1000);border:1px solid var(--accent2);border-radius:16px;padding:24px 28px;margin-bottom:24px;display:none}
.pay-banner h3{color:var(--accent2);font-size:1rem;margin-bottom:8px}
.pay-banner p{color:var(--muted);font-size:.83rem;margin-bottom:16px;line-height:1.6}
.pay-row{display:flex;gap:12px;align-items:center;flex-wrap:wrap}
.price-tag{font-size:1.5rem;font-weight:800;color:var(--accent2)}
.price-tag small{font-size:.8rem;color:var(--muted);font-weight:400}

/* Log terminal */
.terminal{background:#000;border:1px solid #1a2a1a;border-radius:14px;padding:20px;margin-top:24px;display:none}
.terminal-header{display:flex;align-items:center;gap:8px;margin-bottom:16px}
.dot{width:10px;height:10px;border-radius:50%}
.terminal-title{font-size:.75rem;color:var(--muted);font-family:'JetBrains Mono',monospace}
#log-output{font-family:'JetBrains Mono',monospace;font-size:.78rem;line-height:1.8;color:#00ff88;height:280px;overflow-y:auto;white-space:pre-wrap}

/* Stats bar */
.stats-bar{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-top:20px;display:none}
.stat-box{background:var(--surface);border:1px solid var(--border);border-radius:12px;padding:16px;text-align:center}
.stat-num{font-size:1.8rem;font-weight:800;font-family:'JetBrains Mono',monospace}
.stat-label{font-size:.7rem;color:var(--muted);text-transform:uppercase;letter-spacing:.08em;margin-top:4px}
.stat-box.high .stat-num{color:var(--high)}
.stat-box.med .stat-num{color:var(--med)}
.stat-box.low .stat-num{color:var(--low)}

/* Download */
.download-section{margin-top:20px;text-align:center;display:none}

/* Footer */
footer{text-align:center;padding:40px 0 20px;color:var(--muted);font-size:.78rem}
footer a{color:var(--muted)}

/* Minor sub-type grid (smaller cards, single-row) */
.minor-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(160px,1fr));gap:10px;margin-bottom:20px}
.minor-grid .niche-label{font-size:.78rem;padding:10px 13px;border-color:#162030}
.minor-grid .niche-item input:checked + .niche-label{border-color:var(--accent2);background:rgba(255,107,53,.07);color:var(--accent2)}

/* Responsive */
@media(max-width:640px){
  .form-row{grid-template-columns:1fr}
  .pricing{grid-template-columns:1fr}
  .niche-grid{grid-template-columns:1fr 1fr}
  .stats-bar{grid-template-columns:1fr 1fr}
}

/* Animations */
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.5}}
.pulsing{animation:pulse 1.5s infinite}
@keyframes slideIn{from{opacity:0;transform:translateY(10px)}to{opacity:1;transform:translateY(0)}}
.slide-in{animation:slideIn .4s ease forwards}
</style>
</head>
<body>
<div class="orb orb1"></div>
<div class="orb orb2"></div>
<div class="container">

<!-- Nav -->
<nav>
  <div class="logo">Lead<span>Hunter</span> <span class="badge">PRO</span></div>
  <div class="nav-links">
    <a href="#pricing">Pricing</a>
    <a href="#scraper">Scraper</a>
  </div>
</nav>

<!-- Hero -->
<div class="hero">
  <h1>Find Businesses<br><span class="green">Without Websites</span><br><span class="orange">Sell Them One</span></h1>
  <p>Scrape Google Maps across 10 niche categories. Extract leads with no website, phone numbers, ratings & reviews — exported to formatted Excel.</p>
  <div class="plan-pills">
    <span class="pill pill-free">✓ FREE — 2 niches forever</span>
    <span class="pill pill-pro">⚡ PRO — $10 per niche</span>
  </div>
</div>

<!-- Pricing -->
<div id="pricing" class="pricing">
  <div class="price-card">
    <h3>Free Tier</h3>
    <div class="price-amount">$0 <small>/ forever</small></div>
    <p class="price-desc">Perfect to test the tool and grab your first leads.</p>
    <ul class="price-features">
      <li>2 niches: Hospitality & Retail</li>
      <li>All data columns included</li>
      <li>Excel export</li>
      <li class="no">Locked niches</li>
      <li class="no">Priority scraping</li>
    </ul>
  </div>
  <div class="price-card featured">
    <h3>Per Niche</h3>
    <div class="price-amount">$10 <small>/ niche</small></div>
    <p class="price-desc">Unlock any niche. Pay once per session. No subscription required.</p>
    <ul class="price-features">
      <li>Any of 10 niche categories</li>
      <li>Phone + Rating + Reviews</li>
      <li>Lead quality scoring</li>
      <li>Multi-sheet Excel output</li>
      <li>Summary dashboard tab</li>
    </ul>
  </div>
</div>

<!-- Scraper Form -->
<div id="scraper" class="card slide-in">
  <div class="section-label">🎯 Configure Your Scrape</div>

  <div class="form-row">
    <div class="form-group">
      <label>Country</label>
      <input type="text" id="country" placeholder="e.g. United States" />
    </div>
    <div class="form-group">
      <label>City</label>
      <input type="text" id="city" placeholder="e.g. New York" />
    </div>
  </div>

  <!-- STEP 1: Major Category -->
  <div class="section-label">Step 1 — Choose a Major Category</div>
  <div class="niche-grid" id="major-grid">
    {% for name, icon in niches %}
    <div class="niche-item">
      <input type="radio" name="major" id="maj_{{loop.index}}" value="{{name}}"
        {% if name in free_niches %}class="free-niche"{% else %}class="paid-niche"{% endif %}
        onchange="onMajorChange(this)">
      <label class="niche-label" for="maj_{{loop.index}}">
        <span class="icon">{{icon}}</span>
        <span class="name">{{name}}</span>
        {% if name in free_niches %}
          <span class="free-tag">FREE</span>
        {% else %}
          <span class="locked">$10</span>
        {% endif %}
      </label>
    </div>
    {% endfor %}
  </div>

  <!-- STEP 2: Minor Sub-type (revealed after major chosen) -->
  <div id="minor-section" style="display:none">
    <div class="section-label" style="margin-top:24px">Step 2 — Choose ONE Sub-Type</div>
    <div class="minor-info" id="minor-info"></div>
    <div class="minor-grid" id="minor-grid"></div>
  </div>

  <!-- Payment banner -->
  <div class="pay-banner" id="pay-banner">
    <h3>💳 Unlock This Category</h3>
    <p>This is a paid category. Pay $10 to unlock scraping for the selected sub-type.</p>
    <div class="pay-row">
      <span class="price-tag">$10 <small>one-time</small></span>
      <button class="btn btn-pay" onclick="handlePayment()">⚡ Pay & Unlock</button>
    </div>
  </div>

  <button class="btn btn-primary btn-full" id="start-btn" onclick="startScrape()" style="margin-top:20px">
    🚀 Start Scraping
  </button>
</div>

<!-- Terminal log -->
<div class="terminal" id="terminal">
  <div class="terminal-header">
    <div class="dot" style="background:#ef4444"></div>
    <div class="dot" style="background:#f59e0b"></div>
    <div class="dot" style="background:#22c55e"></div>
    <span class="terminal-title">LeadHunter — Live Output</span>
  </div>
  <div id="log-output"></div>
</div>

<!-- Stats -->
<div class="stats-bar" id="stats-bar">
  <div class="stat-box"><div class="stat-num" id="s-total">0</div><div class="stat-label">Total Leads</div></div>
  <div class="stat-box high"><div class="stat-num" id="s-high">0</div><div class="stat-label">High Leads</div></div>
  <div class="stat-box med"><div class="stat-num" id="s-med">0</div><div class="stat-label">Medium</div></div>
  <div class="stat-box low"><div class="stat-num" id="s-low">0</div><div class="stat-label">Low</div></div>
</div>

<!-- Download -->
<div class="download-section" id="download-section">
  <button class="btn btn-primary" id="dl-btn" style="font-size:1rem;padding:16px 36px">
    📥 Download Excel File
  </button>
</div>

<footer>
  <p>⚠️ LeadHunter is for educational use. Google Maps scraping may violate ToS. Use responsibly.</p>
  <p style="margin-top:8px"><a href="https://policies.google.com/terms" target="_blank">Google ToS</a> · <a href="#">Privacy</a> · Built with ❤️</p>
</footer>

</div>

<script>
// ── Data passed from Flask ────────────────────────────────────────
const NICHES_DATA  = {{ niches_data | tojson }};   // {name: [sub1, sub2, ...]}
const FREE_NICHES  = {{ free_niches | tojson }};
let   paidNiches   = {{ paid_niches | tojson }};
let   currentJobId = null;

// ── Step 1: user picks a major category ──────────────────────────
function onMajorChange(radio) {
  const majorName = radio.value;
  const subs      = NICHES_DATA[majorName] || [];
  const isPaid    = !FREE_NICHES.includes(majorName) && !paidNiches.includes(majorName);

  // Build minor grid
  const grid = document.getElementById('minor-grid');
  grid.innerHTML = '';
  subs.forEach((sub, i) => {
    grid.innerHTML += `
      <div class="niche-item">
        <input type="radio" name="minor" id="min_${i}" value="${sub}" onchange="onMinorChange()">
        <label class="niche-label" for="min_${i}">
          <span class="icon">🔹</span>
          <span class="name">${sub}</span>
        </label>
      </div>`;
  });

  document.getElementById('minor-info').innerHTML =
    `<p style="color:var(--muted);font-size:.82rem;margin-bottom:14px">
      Showing ${subs.length} sub-types under <strong style="color:var(--text)">${majorName}</strong>.
      Pick one to scrape — fast &amp; focused.
    </p>`;

  document.getElementById('minor-section').style.display = 'block';
  document.getElementById('pay-banner').style.display    = isPaid ? 'block' : 'none';
}

// ── Step 2: user picks a minor sub-type ──────────────────────────
function onMinorChange() {
  // Nothing extra needed — start button is always visible
}

// ── Payment ───────────────────────────────────────────────────────
async function handlePayment() {
  const majorEl = document.querySelector('input[name="major"]:checked');
  if (!majorEl) return;
  const r = await fetch('/demo-pay', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({niches: [majorEl.value]})
  });
  const data = await r.json();
  if (data.success) {
    paidNiches = data.paid_niches;
    document.getElementById('pay-banner').style.display = 'none';
    alert('✅ Unlocked! You can now scrape ' + majorEl.value);
  }
}

// ── Start scrape ──────────────────────────────────────────────────
async function startScrape() {
  const country  = document.getElementById('country').value.trim();
  const city     = document.getElementById('city').value.trim();
  const majorEl  = document.querySelector('input[name="major"]:checked');
  const minorEl  = document.querySelector('input[name="minor"]:checked');

  if (!country || !city)  { alert('Please enter country and city.'); return; }
  if (!majorEl)           { alert('Please select a major category (Step 1).'); return; }
  if (!minorEl)           { alert('Please select a sub-type (Step 2).'); return; }

  const major = majorEl.value;
  const minor = minorEl.value;
  const isPaid = !FREE_NICHES.includes(major) && !paidNiches.includes(major);
  if (isPaid) { alert('Please complete payment to unlock this category.'); return; }

  document.getElementById('start-btn').disabled = true;
  document.getElementById('start-btn').textContent = '⏳ Scraping…';
  document.getElementById('terminal').style.display = 'block';
  document.getElementById('log-output').textContent = '';
  document.getElementById('stats-bar').style.display = 'none';
  document.getElementById('download-section').style.display = 'none';

  const r = await fetch('/start', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({country, city, major, minor})
  });
  const data = await r.json();
  currentJobId = data.job_id;
  listenLogs(currentJobId);
}

function listenLogs(jobId) {
  const out = document.getElementById('log-output');
  const es  = new EventSource('/logs/' + jobId);
  es.onmessage = e => {
    out.textContent += e.data + '\n';
    out.scrollTop = out.scrollHeight;
    if (e.data.includes('DONE')) {
      es.close();
      fetchStats(jobId);
      const btn = document.getElementById('start-btn');
      btn.disabled = false;
      btn.textContent = '🚀 Start Scraping';
    }
  };
  es.onerror = () => es.close();
}

async function fetchStats(jobId) {
  const r = await fetch('/stats/' + jobId);
  const d = await r.json();
  document.getElementById('s-total').textContent = d.total;
  document.getElementById('s-high').textContent  = d.high;
  document.getElementById('s-med').textContent   = d.med;
  document.getElementById('s-low').textContent   = d.low;
  document.getElementById('stats-bar').style.display = 'grid';
  document.getElementById('dl-btn').onclick = () => window.location='/download/'+jobId;
  document.getElementById('download-section').style.display = 'block';
}
</script>
</body>
</html>"""


@app.route("/")
def index():
    if "session_id" not in session:
        session["session_id"] = str(uuid.uuid4())
    paid = list(session.get("paid_niches", []))
    niches = [(name, NICHE_ICONS.get(name, "📌")) for name in NICHES]
    niches_data = {name: subs for name, subs in NICHES.items()}
    return render_template_string(HTML,
                                  niches=niches,
                                  niches_data=niches_data,
                                  free_niches=FREE_NICHES,
                                  paid_niches=paid,
                                  stripe_pub_key=STRIPE_PUB_KEY)


@app.route("/demo-pay", methods=["POST"])
def demo_pay():
    """Demo payment endpoint — replace with real Stripe in production."""
    data = request.json
    niches = data.get("niches", [])
    paid = set(session.get("paid_niches", []))
    paid.update(niches)
    session["paid_niches"] = list(paid)
    return jsonify({"success": True, "paid_niches": list(paid)})


@app.route("/create-checkout-session", methods=["POST"])
def create_checkout():
    """Real Stripe checkout — activate when you have Stripe keys."""
    data = request.json
    niches = data.get("niches", [])
    count = len(niches)
    try:
        checkout = stripe.checkout.Session.create(
            payment_method_types=["card"],
            line_items=[{
                "price_data": {
                    "currency": "usd",
                    "product_data": {"name": f"LeadHunter Pro — {count} niche(s)"},
                    "unit_amount": PRICE_PER_NICHE * count,
                },
                "quantity": 1,
            }],
            mode="payment",
            success_url=request.host_url + f"pay-success?niches={','.join(niches)}",
            cancel_url=request.host_url,
        )
        return jsonify({"checkout_url": checkout.url})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/pay-success")
def pay_success():
    niches = request.args.get("niches", "").split(",")
    paid = set(session.get("paid_niches", []))
    paid.update(niches)
    session["paid_niches"] = list(paid)
    return redirect(url_for("index"))


@app.route("/start", methods=["POST"])
def start_job():
    data = request.json
    country = data.get("country", "")
    city = data.get("city", "")
    major = data.get("major", "")
    minor = data.get("minor", "")
    job_id = str(uuid.uuid4())

    job_queues[job_id] = Queue()
    job_results[job_id] = []
    job_status[job_id] = "running"

    def run():
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        # Pass single-item list so run_scrape iterates once
        loop.run_until_complete(run_scrape(city, country, [major], job_id,
                                           single_minor=minor))
        loop.close()

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/logs/<job_id>")
def stream_logs(job_id):
    def generate():
        q = job_queues.get(job_id)
        if not q:
            return
        while True:
            try:
                msg = q.get(timeout=30)
                yield f"data: {msg}\n\n"
                if "DONE" in msg:
                    break
            except Empty:
                yield "data: ⏳ Still working…\n\n"

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/stats/<job_id>")
def get_stats(job_id):
    records = job_results.get(job_id, [])
    return jsonify({
        "total": len(records),
        "high": sum(1 for r in records if r.get("Lead Quality") == "HIGH LEAD"),
        "med": sum(1 for r in records if r.get("Lead Quality") == "MEDIUM LEAD"),
        "low": sum(1 for r in records if r.get("Lead Quality") == "LOW LEAD"),
    })


@app.route("/download/<job_id>")
def download(job_id):
    records = job_results.get(job_id, [])
    # Extract city/country from first record's address or use generic name
    path = export_excel(records, "export", "leads")
    return send_file(str(path), as_attachment=True,
                     download_name=path.name,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=5000, threaded=True)